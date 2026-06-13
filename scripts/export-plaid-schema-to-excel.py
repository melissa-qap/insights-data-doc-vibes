#!/usr/bin/env python3
"""Generate plaid-api-schema.xlsx from plaid-api-schema.md."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SCHEMA_MD = ROOT / ".cursor/skills/outline-plaid-insight/plaid-api-schema.md"
OUTPUT_XLSX = ROOT / ".cursor/skills/outline-plaid-insight/plaid-api-schema.xlsx"

# Comment out entries to exclude sheets from export.
SHEETS = [
    # Plaid datatables
    "plaid_items",
    "plaid_accounts",
    "plaid_transactions",
    "plaid_investment_holdings",
    "plaid_investment_securities",
    "plaid_liabilities_credit",
    "plaid_liabilities_student",
    "plaid_liabilities_mortgage",
    # Enum lookup tables
    "enum_account_type",
    "enum_account_subtype",
    "enum_payment_channel",
    "enum_pfc_primary",
    "enum_pfc_detailed",
    "enum_pfc_confidence_level",
    "enum_investment_security_type",
    "enum_investment_security_subtype",
    "enum_loan_status_type",
    "enum_interest_rate_type",
]

DATATABLES = {
    name for name in SHEETS if name.startswith("plaid_")
}

COMMON_COLUMNS = [
    ["`user_id`", "string", "User scope — filter all queries by this"],
    ["`item_id`", "string", "Plaid Item that produced this row"],
    ["`synced_at`", "timestamp", "When this row was imported from Plaid"],
]

SECTION_HEADING = re.compile(r"^### `(?P<name>[^`]+)`\s*$", re.MULTILINE)
SECTION_HEADING_PLAIN = re.compile(r"^### (?P<name>enum_[a-z_]+)\s*$", re.MULTILINE)
PIPE_PLACEHOLDER = "\x00"
EXCEL_SHEET_NAME_MAX = 31


def split_table_row(line: str) -> list[str]:
    """Split a markdown table row, respecting escaped pipes (\\|)."""
    inner = line.strip().strip("|")
    escaped = inner.replace("\\|", PIPE_PLACEHOLDER)
    return [cell.strip().replace(PIPE_PLACEHOLDER, "|") for cell in escaped.split("|")]


def excel_sheet_name(name: str) -> str:
    if len(name) <= EXCEL_SHEET_NAME_MAX:
        return name
    return name[:EXCEL_SHEET_NAME_MAX]


def load_schema_text() -> str:
    text = SCHEMA_MD.read_text(encoding="utf-8")
    api_ref = text.find("## API Field Reference")
    if api_ref != -1:
        text = text[:api_ref]
    return text


def parse_sections(text: str) -> dict[str, str]:
    """Return section name -> body text (content until next ### or ## heading)."""
    sections: dict[str, str] = {}
    matches: list[tuple[int, str]] = []

    for match in SECTION_HEADING.finditer(text):
        matches.append((match.start(), match.group("name")))
    for match in SECTION_HEADING_PLAIN.finditer(text):
        matches.append((match.start(), match.group("name")))

    matches.sort(key=lambda item: item[0])

    for index, (start, name) in enumerate(matches):
        heading_end = text.find("\n", start) + 1
        if index + 1 < len(matches):
            end = matches[index + 1][0]
        else:
            end = len(text)
        sections[name] = text[heading_end:end].strip()

    return sections


def parse_markdown_table(body: str) -> tuple[list[str], list[list[str]]]:
    """Extract the first markdown pipe table from section body."""
    lines = body.splitlines()
    table_lines: list[str] = []
    in_table = False

    for line in lines:
        stripped = line.strip()
        if stripped.startswith("|") and stripped.endswith("|"):
            in_table = True
            table_lines.append(stripped)
        elif in_table:
            break

    if not table_lines:
        raise ValueError("No markdown table found in section")

    rows: list[list[str]] = []
    for line in table_lines:
        cells = split_table_row(line)
        if all(re.fullmatch(r"-+", cell.replace(" ", "")) or cell == "" for cell in cells):
            continue
        rows.append(cells)

    if not rows:
        raise ValueError("Markdown table has no data rows")

    header, *data = rows
    return header, data


def extract_source_note(body: str) -> str | None:
    for line in body.splitlines():
        stripped = line.strip()
        if stripped.startswith("Source:"):
            return stripped
    return None


def extract_enum_applies_to(body: str) -> str | None:
    for line in body.splitlines():
        stripped = line.strip()
        if stripped.startswith("`") and "—" in stripped:
            return stripped
    return None


def normalize_third_column_header(header: list[str]) -> list[str]:
    if len(header) == 3 and header[2] in ("Notes", "Source field"):
        return [header[0], header[1], "Source field / Notes"]
    return header


def prepend_common_columns(header: list[str], data: list[list[str]]) -> tuple[list[str], list[list[str]]]:
    normalized_header = normalize_third_column_header(header)
    common_header = normalized_header
    merged_data = [row[: len(common_header)] for row in COMMON_COLUMNS]
    merged_data.extend(data)
    return common_header, merged_data


def autofit_columns(ws, max_width: int = 60) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)


def write_sheet(
    wb: Workbook,
    sheet_name: str,
    title: str,
    subtitle: str | None,
    header: list[str],
    data: list[list[str]],
) -> None:
    ws = wb.create_sheet(title=excel_sheet_name(sheet_name))
    row = 1
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    if subtitle:
        ws.cell(row=row, column=1, value=subtitle)
        row += 1

    header_row = row
    for col_idx, value in enumerate(header, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=value)
        cell.font = Font(bold=True)

    for offset, data_row in enumerate(data, start=1):
        for col_idx, value in enumerate(data_row, start=1):
            ws.cell(row=header_row + offset, column=col_idx, value=value)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    autofit_columns(ws)


def build_workbook(sections: dict[str, str]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    for name in SHEETS:
        if name not in sections:
            raise KeyError(f"Section not found in schema: {name}")

        body = sections[name]
        header, data = parse_markdown_table(body)

        if name in DATATABLES and name != "plaid_items":
            header, data = prepend_common_columns(header, data)
            subtitle = extract_source_note(body)
        elif name.startswith("enum_"):
            subtitle = extract_enum_applies_to(body)
        else:
            subtitle = extract_source_note(body)

        write_sheet(
            wb=wb,
            sheet_name=name,
            title=name,
            subtitle=subtitle,
            header=header,
            data=data,
        )

    return wb


def main() -> None:
    text = load_schema_text()
    sections = parse_sections(text)
    wb = build_workbook(sections)
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"Wrote {len(SHEETS)} sheets to {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
